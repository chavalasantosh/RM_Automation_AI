"""
Excel Report Generator for RME
Creates detailed match analysis reports in Excel format with interactive features.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.utils import get_column_letter
import json
from typing import Dict, List, Any, Optional

class ExcelReport:
    """Generates detailed Excel reports for match analysis."""
    
    def __init__(self, match_results: List[Dict], output_dir: str = "output"):
        """
        Initialize the Excel report generator.
        
        Args:
            match_results: List of match results from the main system
            output_dir: Directory to save the Excel file
        """
        self.match_results = match_results
        self.output_dir = output_dir
        self.wb = Workbook()
        self._setup_styles()
        
    def _setup_styles(self):
        """Set up common styles for the Excel report."""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_aligned = Alignment(horizontal='center', vertical='center')
        
    def _create_summary_sheet(self):
        """Create the executive summary sheet."""
        ws = self.wb.active
        ws.title = "Executive Summary"
        
        # Add title
        ws['A1'] = "Resume Matching Analysis Report"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Add timestamp
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:F2')
        
        # Summary statistics
        stats = {
            "Total Matches": len(self.match_results),
            "Perfect Matches (>80%)": sum(1 for r in self.match_results if r['match_score'] > 80),
            "Average Match Score": f"{sum(r['match_score'] for r in self.match_results) / len(self.match_results):.1f}%",
            "Highest Match": f"{max(r['match_score'] for r in self.match_results):.1f}%",
            "Lowest Match": f"{min(r['match_score'] for r in self.match_results):.1f}%"
        }
        
        # Add statistics
        for i, (key, value) in enumerate(stats.items(), start=4):
            ws[f'A{i}'] = key
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            
        # Add match overview table
        headers = ["Job", "Candidate", "Match Score", "Status", "Experience Match", "Certification Match"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=10, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_aligned
            
        # Add match data
        for i, result in enumerate(self.match_results, start=11):
            status = 'Perfect Match' if result['match_score'] > 80 else 'Good Match' if result['match_score'] > 60 else 'Moderate Match'
            data = [
                result['job'],
                result['profile'],
                f"{result['match_score']}%",
                status,
                f"{result['matching_criteria']['experience_match']}%",
                f"{result['matching_criteria']['certification_match']}%"
            ]
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=i, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = self.center_aligned
                
        # Add match score chart
        chart = BarChart()
        chart.title = "Match Scores Overview"
        chart.style = 10
        chart.y_axis.title = "Match Score (%)"
        chart.x_axis.title = "Job-Candidate Pairs"
        
        data = Reference(ws, min_col=3, min_row=10, max_row=10+len(self.match_results), max_col=3)
        cats = Reference(ws, min_col=1, min_row=11, max_row=10+len(self.match_results), max_col=2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "A20")
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
            
    def _create_skill_analysis_sheet(self):
        """Create the skill analysis sheet with detailed skill matching."""
        ws = self.wb.create_sheet("Skill Analysis")
        
        # Add title
        ws['A1'] = "Detailed Skill Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Create skill matrix
        headers = ["Job", "Candidate", "Required Skills", "Matching Skills", "Missing Skills"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_aligned
            
        # Add skill data
        row = 4
        for result in self.match_results:
            job_data = result['job_data']
            profile_data = result['profile_data']
            
            required_skills = set(job_data.get('required_skills', []))
            profile_skills = set(profile_data.get('skills', []))
            matching_skills = required_skills.intersection(profile_skills)
            missing_skills = required_skills - profile_skills
            
            data = [
                result['job'],
                result['profile'],
                ', '.join(required_skills),
                ', '.join(matching_skills),
                ', '.join(missing_skills)
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                
            row += 1
            
        # Adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 30
            
    def _create_candidate_details_sheet(self):
        """Create detailed candidate analysis sheet."""
        ws = self.wb.create_sheet("Candidate Details")
        
        # Add title
        ws['A1'] = "Detailed Candidate Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Create candidate details table
        headers = ["Candidate", "Education", "Experience", "Certifications", "Skills", "Match Score"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_aligned
            
        # Add candidate data
        row = 4
        for result in self.match_results:
            profile_data = result['profile_data']
            data = [
                result['profile'],
                profile_data.get('education', 'N/A'),
                f"{profile_data.get('years_of_experience', 0)} years",
                ', '.join(profile_data.get('certifications', [])),
                ', '.join(profile_data.get('skills', [])),
                f"{result['match_score']}%"
            ]
            
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                
            row += 1
            
        # Add radar chart for skills
        chart = RadarChart()
        chart.type = "radar"
        chart.style = 26
        chart.title = "Skill Match Analysis"
        
        # Add chart data
        # (This would be implemented based on the actual skill matching data)
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 25
            
    def generate(self, filename: Optional[str] = None) -> str:
        """
        Generate the Excel report.
        
        Args:
            filename: Optional custom filename for the report
            
        Returns:
            str: Path to the generated Excel file
        """
        if filename is None:
            filename = f"detailed_match_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
            
        # Create all sheets
        self._create_summary_sheet()
        self._create_skill_analysis_sheet()
        self._create_candidate_details_sheet()
        
        # Save the workbook
        output_path = os.path.join(self.output_dir, filename)
        self.wb.save(output_path)
        
        return output_path

def generate_excel_report(match_results: List[Dict], output_dir: str = "output") -> str:
    """
    Convenience function to generate an Excel report.
    
    Args:
        match_results: List of match results from the main system
        output_dir: Directory to save the Excel file
        
    Returns:
        str: Path to the generated Excel file
    """
    report = ExcelReport(match_results, output_dir)
    return report.generate()

if __name__ == "__main__":
    # Example usage
    with open('output/match_results.txt', 'r') as f:
        # This is just an example - in reality, you'd use the actual match results
        sample_results = [
            {
                'job': 'Senior Developer',
                'profile': 'John Smith',
                'match_score': 95.0,
                'matching_criteria': {
                    'experience_match': 100.0,
                    'certification_match': 90.0
                },
                'job_data': {
                    'required_skills': ['Python', 'Django', 'SQL']
                },
                'profile_data': {
                    'education': 'Master\'s in CS',
                    'years_of_experience': 7,
                    'certifications': ['AWS Certified'],
                    'skills': ['Python', 'Django', 'SQL', 'Git']
                }
            }
        ]
    
    output_file = generate_excel_report(sample_results)
    print(f"Excel report generated: {output_file}") 